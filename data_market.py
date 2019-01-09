# -*- coding: utf-8 -*-
"""
Created on Mon Jan  7 10:40:00 2019

@author: RUFIAR1
"""
from openpyxl import load_workbook
import pandas as pd
import ai_analysis.create_data as cd

def create_market_data_complete(file='Market Data/2018.10 RETAIL - Download 20 sample molecule - 2018.12.20 - Copy.xlsx',sheet_name='2018.10 RETAIL - Download 20 sa'):    
    wb = load_workbook(filename = file)
    intrested_sheet=wb[sheet_name]
    market_data=cd.create_market_data(intrested_sheet,0,intrested_sheet.max_row)
    market_data['All Data']=market_data['Molecule']+','+market_data['Manufacturer']+','+market_data['Product']+','+market_data['Pack']+','+market_data['BRAND-INN']+','+market_data["GX-OX"]
    market_data=market_data.drop(columns=['Molecule','Molecule ADJ','Manufacturer','Product','Pack','BRAND-INN','GX-OX'])
    market_data=market_data.set_index("All Data")
    return market_data
    
wb=load_workbook(filename='Market Data/2018.10 RETAIL - Download 20 sample molecule - 2018.12.20 - Copy.xlsx')
intrested_sheet=wb['2018.10 RETAIL - Download 20 sa']
market_data=cd.create_market_data(intrested_sheet,0,intrested_sheet.max_row)
market_data['All Data']=market_data['Molecule']+','+market_data['Manufacturer']+','+market_data['Product']+','+market_data['Pack']+','+market_data['BRAND-INN']+','+market_data["GX-OX"]
market_data=market_data.drop(columns=['Molecule','Molecule ADJ','Manufacturer','Product','Pack','BRAND-INN','GX-OX'])
market_data=market_data.set_index("All Data").T

m=create_market_data_complete()
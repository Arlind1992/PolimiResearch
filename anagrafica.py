# -*- coding: utf-8 -*-
"""
Created on Mon Jan  7 15:47:51 2019

@author: RUFIAR1
"""
from openpyxl import load_workbook
import pandas as pd

def create_anagrafica(file='Supply&Demand/anagrafica_AI.xlsx',sheet_name='report Bi'):
    wb=load_workbook(filename=file)
    sheet=wb[sheet_name]
    to_return={}
    index_row=sheet[2]
    for index_cells in index_row:
        if(index_cells.value):
            to_return[index_cells.value]=[sheet[index_cells.column+str(x)].value or 0 for x in range(3,sheet.max_row)]
    return pd.DataFrame.from_records(to_return)

 
a=create_anagrafica()

a=a.sort_values(by='Material')
'''perche sono attivi?'''
subset=a[a.duplicated(subset=['GMD FDF ID','Brand'])]

subt=subset[(subset['Material']==44062070) | (subset['Material']==44067807) | (subset['Material']==44070378) | (subset['Material']==44057383) | (subset['Material']==44000735) | (subset['Material']==44058951) | (subset['Material']==44012619 )|  (subset['Material']==44060859)]
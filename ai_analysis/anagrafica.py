# -*- coding: utf-8 -*-
"""
Created on Mon Jan  7 15:47:51 2019

@author: RUFIAR1
"""
from openpyxl import load_workbook
import pandas as pd

def create_anagrafica(index_row_num,file='Supply&Demand/anagrafica_AI.xlsx',sheet_name='report Bi'):
    wb=load_workbook(filename=file)
    sheet=wb[sheet_name]
    to_return={}
    index_row=sheet[index_row_num]
    for index_cells in index_row:
        if(index_cells.value):
            to_return[index_cells.value]=[sheet[index_cells.column+str(x)].value or 0 for x in range(index_row_num+1,sheet.max_row)]
    return pd.DataFrame.from_records(to_return)


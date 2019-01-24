#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Jan 22 12:36:00 2019

@author: arlind
"""

from openpyxl import load_workbook
import ai_analysis.manual_integration.create_data as cd

def get_sales_data():
    sales_workbook=load_workbook(filename='AllData/allSales.xlsm')
    sales_sheet=sales_workbook['Table']
    sales_data=cd.create_finance_data(sales_sheet,14,sales_sheet.max_row)
    list_columns=list(sales_data.columns.values)
    for col_name in list_columns:
        if not ('20' in col_name or 'Material' in col_name):
            sales_data=sales_data.drop(col_name,axis=1)
    return sales_data

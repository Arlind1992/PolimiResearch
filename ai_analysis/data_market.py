# -*- coding: utf-8 -*-
"""
Created on Mon Jan  7 10:40:00 2019

@author: RUFIAR1
"""
from openpyxl import load_workbook
import pandas as pd
import ai_analysis.manual_integration.create_data as cd
import re
from os import listdir
from os.path import isfile, join
def create_market_data_complete(file='Market Data/2018.10 RETAIL - Download 20 sample molecule - 2018.12.20 - Copy.xlsx',sheet_name='2018.10 RETAIL - Download 20 sa'):    
    wb = load_workbook(filename = file)
    intrested_sheet=wb[sheet_name]
    market_data=cd.create_market_data(intrested_sheet,0,intrested_sheet.max_row)
    market_data['All Data']=market_data['Molecule']+','+market_data['Manufacturer']+','+market_data['Product']+','+market_data['Pack']+','+market_data['BRAND-INN']+','+market_data["GX-OX"]
    market_data=market_data.drop(columns=['Molecule','Molecule ADJ','Manufacturer','Product','Pack','BRAND-INN','GX-OX'])
    market_data=market_data.set_index("All Data")
    return market_data

def create_market_data_from_csv(filepath='Market Data/All Market Data Sandoz 2.csv',separator=',',multiply=1,decimalsep=','):
    dt=pd.read_csv(filepath,sep=separator,decimal=decimalsep) 
    col_names=list(dt)
    pattern=re.compile('^Sell-in (.*?)/')
    for colname in col_names:
        if pattern.match(colname):
            dt[colname]=dt[colname].apply(lambda x: int(str(x).replace('.','')) if multiply==1 else int(str(x).replace('.',''))*multiply).astype(int)
            dt=dt.rename(columns = {colname: re.sub(pattern, '01/', colname)})
    return dt

def create_all_old_market_data(path_data):
    onlyfiles = [f for f in listdir(path_data) if isfile(join(path_data, f))]
    to_multiply=1000 if 'thousend' in onlyfiles[0] else 1
    df=create_market_data_from_csv(filepath=path_data+'/'+onlyfiles[0],separator=';',multiply=to_multiply)
    for f_pos in range(1,len(onlyfiles)):
        to_multiply=1000 if 'thousend' in onlyfiles[f_pos] else 1
        df=df.merge(create_market_data_from_csv(filepath=path_data+'/'+onlyfiles[f_pos],separator=';',multiply=to_multiply), how='outer',on=['Manufacturer','Name Type','Product','Pack','Molecule','Anatomical Therapeutic Class 4'] )    
    return df

def add_latest_data(latest_data_path,old_data_path):
    latest_data=create_market_data_from_csv(filepath=latest_data_path,separator=';')
    df=create_market_data_from_csv(filepath=old_data_path, separator=';')
    for col_name in list(latest_data.columns.values):
        if '01' in col_name:
            try:
                df=df.drop(columns=col_name, axis=1)
            except:
                pass
    toreturn=df.merge(latest_data, how='outer',on=['Manufacturer','Name Type','Product','Pack','Molecule','Anatomical Therapeutic Class 4'] )    
    return toreturn    

def get_market_data():
    return create_market_data_from_csv(filepath='AllData/AllDataIMS.csv',separator=';').drop(columns='Unnamed: 0.1',axis=1)

def get_probiotici(file='AllData/Probiotici.xlsx',sheet_name='Sheet1'):
    wb = load_workbook(filename = file)
    intrested_sheet=wb[sheet_name]
    market_data_probiotici=cd.create_market_data(intrested_sheet,0,intrested_sheet.max_row).drop(columns='MonthYear',axis=1)
    market_data_probiotici=market_data_probiotici[(market_data_probiotici['Company']!='Total') &(market_data_probiotici['Product']!='Total')&(market_data_probiotici['Brand']!='Total')]
    return market_data_probiotici.replace('-',0)


def get_probiotici_csv(file='AllData/Probiotici.csv'):
    market_data_probiotici=pd.read_csv(filepath_or_buffer=file,sep=';',decimal=',').drop(columns='MonthYear',axis=1)
    market_data_probiotici=market_data_probiotici[(market_data_probiotici['Company']!='Total') &(market_data_probiotici['Product']!='Total')&(market_data_probiotici['Brand']!='Total')]
    col_names=list(market_data_probiotici)
    pattern=re.compile('.*-[0-9]{4}')
    for colname in col_names:
        if pattern.match(colname):
            market_data_probiotici[colname]=market_data_probiotici[colname].apply(lambda x: float(str(x).replace('.','').replace(',','.'))*1000 if x!='-' else 0).astype(int)
    return market_data_probiotici
def remove_dupplicates(market_data):
    columns_list=market_data.columns.values.tolist()
    columns_list.remove('Molecule')
    return market_data.groupby(columns_list,as_index=False).agg(lambda x: ','.join(x))

def get_market_competitor_data_by_material(material,market_data,integration,market_data_perimeter):
    integration_filtered=integration[integration['Material'].astype(str)==str(material)]
    market_data_perimeter_filtered=market_data_perimeter[market_data_perimeter['Key']==(integration_filtered['Product'].iloc[0]+' '+integration_filtered['Pack'].iloc[0])]
    if str(market_data_perimeter_filtered['Special Market'].iloc[0])!='nan':
       perimeter_to_join_by=market_data_perimeter[market_data_perimeter['Special Market']==market_data_perimeter_filtered['Special Market'].iloc[0]]
    else:
       perimeter_to_join_by=market_data_perimeter[market_data_perimeter['Mkt Molecola']==market_data_perimeter_filtered['Mkt Molecola'].iloc[0]] 
    perimeter_to_join_by_only_key=perimeter_to_join_by['Key'].to_frame()
    market_data['Key']=market_data['Product']+' '+market_data['Pack']
    return market_data.merge(perimeter_to_join_by_only_key,on='Key')    

'''
filepath='Market Data/All old data'+'/'+onlyfiles[1]
path_data='Market Data/All old data'
code to save all merge and save all market data archives and also add to them the latest data
df=create_all_old_market_data('Market Data/All old data')
df.to_csv(path_or_buf ='Market Data/AllOldData.csv',sep=';')
toreturn=add_latest_data('Market Data/laterstdataallmolecules.csv','Market Data/AllOldData.csv')
toreturn.to_csv(path_or_buf ='Market Data/AllDataIMS.csv',sep=';')
'''


#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Mon Mar 18 13:59:59 2019

@author: arlind
"""

from openpyxl import load_workbook
import pandas as pd
import ai_analysis.manual_integration.create_data as cd
import numpy as np
import datetime
import ai_analysis.data_loading.load_data_locally as ldl
wb=load_workbook(filename='Supply&Demand/estrazione consensus per AI.xlsx')
sheet=wb['Sheet1']
wb2=load_workbook(filename='Sales/ManualForecast.xlsx')
sheet2=wb2['Copied']
man_forecast=cd.create_sales_data(sheet2,5,sheet.max_row)
apo=pd.read_excel('hosp_retail_percent_num.xlsx')
apoweights=apo[['Material','APO Peso Retail']]
apoweights.index=apoweights['Material'].astype(str)
apoweights=apoweights.drop(columns='Material')
anagrafica=ldl.load_anagrafica()
apo_forecast=cd.create_finance_data(sheet,0,sheet.max_row).drop(columns=['Cal. year / month'])
apo_forecast_fil=apo_forecast[apo_forecast['Type']=='Cons Selling FC'].drop(columns='Type')
apo_forecast_fil.index=apo_forecast_fil['Material']
apo_forecast_fil=apo_forecast_fil.drop(columns='Material')
apo_forecast_retail=apo_forecast_fil.join(apoweights)
apo_forecast_retail=apo_forecast_retail[apo_forecast_retail['APO Peso Retail'].notnull()]
apo_forecast_retail=apo_forecast_retail.mul(apo_forecast_retail['APO Peso Retail'],axis=0).drop(columns=['APO Peso Retail'])
apo_forecast_retail=apo_forecast_retail.rename(columns={'Overall Result':'TOTAL 2018'})
man_forecast_fil=man_forecast[['Product','material','CAUSALE','jan','feb','mar','apr','may','jun','jul','aug','sep','oct','nov','dec']]
man_forecast_2018=man_forecast_fil[(man_forecast_fil['CAUSALE']=='Sales History 2018')|(man_forecast_fil['CAUSALE']=='Forecast 2018')]
man_forecast_2018['Total']=man_forecast_2018['jan'].astype(float)+man_forecast_2018['feb'].astype(float)+man_forecast_2018['mar'].astype(float)+man_forecast_2018['apr'].astype(float)+man_forecast_2018['may'].astype(float)+man_forecast_2018['jun'].astype(float)+man_forecast_2018['jul'].astype(float)+man_forecast_2018['aug'].astype(float)+man_forecast_2018['sep'].astype(float)+man_forecast_2018['oct'].astype(float)+man_forecast_2018['nov'].astype(float)+man_forecast_2018['dec'].astype(float)
man_forecast_2018.index=man_forecast_2018['Product'].astype(str)
man_forecast_2018=man_forecast_2018.drop(columns=['Product','material'])
man_forecast_2018=man_forecast_2018.astype({'jan':float,'feb':float,'mar':float,'apr':float,'may':float,'jun':float,'jul':float,'aug':float,'sep':float,'oct':float,'nov':float,'dec':float,'Total':float})
man_forecast_2018_actual=man_forecast_2018[man_forecast_2018['CAUSALE']=='Sales History 2018'].drop(columns=['CAUSALE'])
for column in man_forecast_2018_actual.columns:
    man_forecast_2018_actual=man_forecast_2018_actual.rename(columns={column:column.upper()+' 2018'})

man_difference=apo_forecast_retail-man_forecast_2018_actual
man_percentage=man_difference/man_forecast_2018_actual

automatic_evaluation=pd.read_csv('ARIMA_FORECASTS.csv',sep=';')
aut_eval=automatic_evaluation.drop(columns=['Unnamed: 0'])
aut_eval.index=aut_eval['Material'].astype(str)

aut_eval_actual=aut_eval[aut_eval['Type']=='Actual'].drop(columns=['Type','Material'])
aut_eval_actual['TOTAL 2018']=aut_eval_actual.sum(axis=1)
aut_eval_forecast=aut_eval[aut_eval['Type']=='Forecast'].drop(columns=['Type','Material'])
aut_eval_forecast['TOTAL 2018']=aut_eval_forecast.sum(axis=1)

apo_forecast_only_predicted=apo_forecast_retail.merge(pd.DataFrame(aut_eval_forecast.index),left_index=True,right_on='Material',how='inner')
apo_forecast_only_predicted.index=apo_forecast_only_predicted['Material']
apo_forecast_only_predicted=apo_forecast_only_predicted.drop(columns=['Material'])

datetime.datetime.strptime("feb", '%b').strftime('2018-%m-%d 00:00:00')
columns_to_trasform=apo_forecast_only_predicted.columns.values
for col in columns_to_trasform:
    try:
        apo_forecast_only_predicted=apo_forecast_only_predicted.rename(columns={col:datetime.datetime.strptime(col, '%b 2018').strftime('2018-%m-%d 00:00:00')})
    except:
        pass

material_aut=apo_forecast_only_predicted.index.tolist()
material_man=aut_eval_actual.index.tolist()
main_list = np.setdiff1d(material_man,material_aut)


comp_eval_diffences=(apo_forecast_only_predicted-aut_eval_actual).abs() - (aut_eval_forecast-aut_eval_actual).abs()
comp_eval_percentage=((apo_forecast_only_predicted-aut_eval_actual).abs() - (aut_eval_forecast-aut_eval_actual).abs())/aut_eval_actual
diff_actual_aut_forecast=aut_eval_forecast-aut_eval_actual
diff_actual_man_forecast=apo_forecast_only_predicted-aut_eval_actual
columns=comp_eval_diffences.columns
apo_forecast_only_predicted=apo_forecast_only_predicted[columns]
aut_eval_actual=aut_eval_actual[columns]
aut_eval_forecast=aut_eval_forecast[columns]
for i in range(0,4):
    comp_eval_diffences['Quarter '+str(i+1)]=comp_eval_diffences[comp_eval_diffences.columns[3*i]]+comp_eval_diffences[comp_eval_diffences.columns[3*i+1]]+comp_eval_diffences[comp_eval_diffences.columns[3*i+2]]
    comp_eval_percentage['Quarter '+str(i+1)]=comp_eval_percentage[comp_eval_percentage.columns[3*i]]+comp_eval_percentage[comp_eval_percentage.columns[3*i+1]]+comp_eval_percentage[comp_eval_percentage.columns[3*i+2]]
    diff_actual_aut_forecast['Quarter '+str(i+1)]=diff_actual_aut_forecast[diff_actual_aut_forecast.columns[3*i]]+diff_actual_aut_forecast[diff_actual_aut_forecast.columns[3*i+1]]+diff_actual_aut_forecast[diff_actual_aut_forecast.columns[3*i+2]]
    diff_actual_man_forecast['Quarter '+str(i+1)]=diff_actual_man_forecast[diff_actual_man_forecast.columns[3*i]]+diff_actual_man_forecast[diff_actual_man_forecast.columns[3*i+1]]+diff_actual_man_forecast[diff_actual_man_forecast.columns[3*i+2]]
    apo_forecast_only_predicted['Quarter '+str(i+1)]=apo_forecast_only_predicted[apo_forecast_only_predicted.columns[3*i]]+apo_forecast_only_predicted[apo_forecast_only_predicted.columns[3*i+1]]+apo_forecast_only_predicted[apo_forecast_only_predicted.columns[3*i+2]]
    aut_eval_actual['Quarter '+str(i+1)]=aut_eval_actual[aut_eval_actual.columns[3*i]]+aut_eval_actual[aut_eval_actual.columns[3*i+1]]+aut_eval_actual[aut_eval_actual.columns[3*i+2]]
    aut_eval_forecast['Quarter '+str(i+1)]=aut_eval_forecast[aut_eval_forecast.columns[3*i]]+aut_eval_forecast[aut_eval_forecast.columns[3*i+1]]+aut_eval_forecast[aut_eval_forecast.columns[3*i+2]]


accuracy_apo=1-(apo_forecast_only_predicted-aut_eval_actual).abs()/aut_eval_actual

accuracy_aut=1-(aut_eval_forecast-aut_eval_actual).abs()/aut_eval_actual



to_save=comp_eval_diffences.join(comp_eval_percentage,lsuffix='_APO VS Aut units',rsuffix='_APO VS Aut %').join(diff_actual_aut_forecast,rsuffix='_Aut-Actual').join(diff_actual_man_forecast,rsuffix='_APO-Actual').join(aut_eval_actual,rsuffix='_Actual').join(accuracy_apo,rsuffix='_Accuracy APO').join(accuracy_aut,rsuffix='_Accuracy Models')

tot_models=[x for x in to_save.columns if 'TOTAL' in x]

f_test=to_save[tot_models]

anagrafica=anagrafica[['Material','Descrizione']]
anagrafica.index=anagrafica['Material']
anagrafica=anagrafica.drop(columns=['Material'])

to_save=to_save.join(anagrafica)
to_save.index=to_save.index+' '+to_save['Descrizione']
to_save=to_save.drop(columns=['Descrizione'])
to_save.to_csv('ForecastResults/Results2018APO.csv',sep=';',decimal=',',float_format='%.2f')

comp_eval_diffences.to_csv('ForecastResults/ComparisonDiff2018.csv',sep=';',decimal=',',float_format='%.2f')

comp_eval_percentage.to_csv('ForecastResults/ComparisonPerc2018.csv',sep=';',decimal=',',float_format='%.2f')



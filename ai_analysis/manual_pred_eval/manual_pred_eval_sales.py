#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Mon Feb 25 09:02:50 2019

@author: arlind
"""

from openpyxl import load_workbook
import pandas as pd
import ai_analysis.manual_integration.create_data as cd
import numpy as np
import datetime
import ai_analysis.data_loading.load_data_locally as ldl
wb=load_workbook(filename='Sales/ManualForecast.xlsx')
sheet=wb['Copied']

anagrafica=ldl.load_anagrafica()
man_forecast=cd.create_sales_data(sheet,5,sheet.max_row)
material_prod=man_forecast[['material','Product']].drop_duplicates()
man_forecast_fil=man_forecast[['Product','material','CAUSALE','jan','feb','mar','apr','may','jun','jul','aug','sep','oct','nov','dec']]
man_forecast_2018=man_forecast_fil[(man_forecast_fil['CAUSALE']=='Sales History 2018')|(man_forecast_fil['CAUSALE']=='Forecast 2018')]
man_forecast_2018['Total']=man_forecast_2018['jan'].astype(float)+man_forecast_2018['feb'].astype(float)+man_forecast_2018['mar'].astype(float)+man_forecast_2018['apr'].astype(float)+man_forecast_2018['may'].astype(float)+man_forecast_2018['jun'].astype(float)+man_forecast_2018['jul'].astype(float)+man_forecast_2018['aug'].astype(float)+man_forecast_2018['sep'].astype(float)+man_forecast_2018['oct'].astype(float)+man_forecast_2018['nov'].astype(float)+man_forecast_2018['dec'].astype(float)
man_forecast_2018.index=man_forecast_2018['Product'].astype(str)
man_forecast_2018=man_forecast_2018.drop(columns=['Product','material'])
man_forecast_2018=man_forecast_2018.astype({'jan':float,'feb':float,'mar':float,'apr':float,'may':float,'jun':float,'jul':float,'aug':float,'sep':float,'oct':float,'nov':float,'dec':float,'Total':float})
man_forecast_2018_actual=man_forecast_2018[man_forecast_2018['CAUSALE']=='Sales History 2018'].drop(columns=['CAUSALE'])
man_forecast_2018_forecast=man_forecast_2018[man_forecast_2018['CAUSALE']=='Forecast 2018'].drop(columns=['CAUSALE'])
man_difference=man_forecast_2018_forecast-man_forecast_2018_actual
man_percentage=man_difference/man_forecast_2018_actual

automatic_evaluation=pd.read_csv('ARIMA_FORECASTS.csv',sep=';')
aut_eval=automatic_evaluation.drop(columns=['Unnamed: 0'])
aut_eval.index=aut_eval['Material'].astype(str)

aut_eval_actual=aut_eval[aut_eval['Type']=='Actual'].drop(columns=['Type','Material'])
aut_eval_actual['Total']=aut_eval_actual.sum(axis=1)
aut_eval_forecast=aut_eval[aut_eval['Type']=='Forecast'].drop(columns=['Type','Material'])
aut_eval_forecast['Total']=aut_eval_forecast.sum(axis=1)

man_forecast_only_predicted=man_forecast_2018_forecast.merge(pd.DataFrame(aut_eval_forecast.index),left_index=True,right_on='Material',how='inner')
man_forecast_only_predicted.index=man_forecast_only_predicted['Material']
man_forecast_only_predicted=man_forecast_only_predicted.drop(columns=['Material'])

datetime.datetime.strptime("feb", '%b').strftime('2018-%m-%d 00:00:00')
columns_to_trasform=man_forecast_only_predicted.columns.values
for col in columns_to_trasform:
    try:
        man_forecast_only_predicted=man_forecast_only_predicted.rename(columns={col:datetime.datetime.strptime(col, '%b').strftime('2018-%m-%d 00:00:00')})
    except:
        pass

material_aut=man_forecast_only_predicted.index.tolist()
material_man=aut_eval_actual.index.tolist()
main_list = np.setdiff1d(material_man,material_aut)


comp_eval_diffences=(man_forecast_only_predicted-aut_eval_actual).abs() - (aut_eval_forecast-aut_eval_actual).abs()
comp_eval_percentage=((man_forecast_only_predicted-aut_eval_actual).abs() - (aut_eval_forecast-aut_eval_actual).abs())/aut_eval_actual
diff_actual_aut_forecast=aut_eval_forecast-aut_eval_actual
diff_actual_man_forecast=man_forecast_only_predicted-aut_eval_actual

for i in range(0,4):
    comp_eval_diffences['Quarter '+str(i+1)]=comp_eval_diffences[comp_eval_diffences.columns[3*i]]+comp_eval_diffences[comp_eval_diffences.columns[3*i+1]]+comp_eval_diffences[comp_eval_diffences.columns[3*i+2]]
    comp_eval_percentage['Quarter '+str(i+1)]=comp_eval_percentage[comp_eval_percentage.columns[3*i]]+comp_eval_percentage[comp_eval_percentage.columns[3*i+1]]+comp_eval_percentage[comp_eval_percentage.columns[3*i+2]]
    diff_actual_aut_forecast['Quarter '+str(i+1)]=diff_actual_aut_forecast[diff_actual_aut_forecast.columns[3*i]]+diff_actual_aut_forecast[diff_actual_aut_forecast.columns[3*i+1]]+diff_actual_aut_forecast[diff_actual_aut_forecast.columns[3*i+2]]
    diff_actual_man_forecast['Quarter '+str(i+1)]=diff_actual_man_forecast[diff_actual_man_forecast.columns[3*i]]+diff_actual_man_forecast[diff_actual_man_forecast.columns[3*i+1]]+diff_actual_man_forecast[diff_actual_man_forecast.columns[3*i+2]]

to_save=comp_eval_diffences.join(comp_eval_percentage,lsuffix='_Man VS Aut units',rsuffix='_Man VS Aut %').join(diff_actual_aut_forecast,rsuffix='_Aut-Actual').join(diff_actual_man_forecast,rsuffix='_Man-Actual').join(aut_eval_actual,rsuffix='_Actual')

anagrafica=anagrafica[['Material','Descrizione']]
anagrafica.index=anagrafica['Material']
anagrafica=anagrafica.drop(columns=['Material'])

to_save=to_save.join(anagrafica)
to_save.index=to_save.index+' '+to_save['Descrizione']
to_save=to_save.drop(columns=['Descrizione'])
to_save.to_csv('ForecastResults/Results2018.csv',sep=';',decimal=',',float_format='%.2f')

comp_eval_diffences.to_csv('ForecastResults/ComparisonDiff2018.csv',sep=';',decimal=',',float_format='%.2f')

comp_eval_percentage.to_csv('ForecastResults/ComparisonPerc2018.csv',sep=';',decimal=',',float_format='%.2f')


